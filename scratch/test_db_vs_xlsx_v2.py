import requests
import openpyxl

SUPABASE_URL = 'https://mujqmdmzloizqhglayxe.supabase.co'
SUPABASE_KEY = 'eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6Im11anFtZG16bG9penFoZ2xheXhlIiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc4MjgxNTk0OCwiZXhwIjoyMDk4MzkxOTQ4fQ.G7AmEylxPwm6TWZ3xjCcBhvhfNPYHdj0V08My0A_rEc'
headers = {
    'apikey': SUPABASE_KEY,
    'Authorization': f'Bearer {SUPABASE_KEY}'
}

# Fetch all from round_1_evaluation
url = f"{SUPABASE_URL}/rest/v1/round_1_evaluation?select=id,app_status,review_comments,eval_group"
r = requests.get(url, headers=headers)
db_map = {row['id']: row for row in r.json()}

python_path = r"c:\Users\Dhanush\Music\aviatorsclaude\r1_r2_side_by_side_report_2026-07-13.xlsx"
wb = openpyxl.load_workbook(python_path)
sheet = wb['Analysis']

raw_sheet = wb['Raw Data']
# Let's map row number in 'Analysis' to candidate ID using the 'Raw Data' tab,
# or by reading Column AE (31) Email or Column AG (33) Github or from ID in Raw Data.
# Wait! In 'Analysis' tab, does it write candidate ID anywhere?
# In the table rows of 'Analysis', we wrote:
# idx 0: Rank (index + 1)
# B: Name
# C: Gender
# ...
# AD: Email (column 30)
# Let's map Analysis row to ID using Email!
# Let's fetch raw_submissions to get mapping of Email -> ID
r_raw = requests.get(f"{SUPABASE_URL}/rest/v1/raw_submissions?select=id,email", headers=headers)
email_to_id = {}
for row in r_raw.json():
    email = str(row.get('email') or '').strip().lower()
    if email:
        email_to_id[email] = row['id']

mismatches = 0
for r in range(7, sheet.max_row + 1):
    email_val = sheet.cell(row=r, column=30).value # Column AD is column 30
    if email_val:
        email_key = str(email_val).strip().lower()
        if email_key in email_to_id:
            cand_id = email_to_id[email_key]
            if cand_id in db_map:
                db_row = db_map[cand_id]
                
                review_excel = sheet.cell(row=r, column=36).value # AJ (R1 Review)
                screener_excel = sheet.cell(row=r, column=38).value # AL (Screener)
                status_excel = sheet.cell(row=r, column=39).value # AM (Status)
                
                # Check for R1 Review mismatches
                db_review = db_row.get('review_comments')
                db_screener = db_row.get('eval_group')
                db_status = db_row.get('app_status')
                
                # Normalize values
                def clean(v):
                    v_str = str(v or '').strip()
                    if v_str in ['', 'None', '-', 'NaN']:
                        return ''
                    return v_str.lower()
                    
                if clean(review_excel) != clean(db_review):
                    mismatches += 1
                    print(f"Mismatch R1 Review: ID={cand_id}, Excel='{review_excel}', DB='{db_review}'")
                if clean(screener_excel) != clean(db_screener):
                    mismatches += 1
                    print(f"Mismatch Screener: ID={cand_id}, Excel='{screener_excel}', DB='{db_screener}'")
                if clean(status_excel) != clean(db_status):
                    mismatches += 1
                    print(f"Mismatch Status: ID={cand_id}, Excel='{status_excel}', DB='{db_status}'")
            else:
                # Missing from round_1_evaluation
                pass

print(f"\nTotal logical data mismatches: {mismatches}")
