import openpyxl

wb = openpyxl.load_workbook(r"c:\Users\Dhanush\Music\aviatorsclaude\Untitled spreadsheet (1).xlsx")
ws = wb.active

row_vals = [ws.cell(row=13, column=c).value for c in range(1, ws.max_column + 1)]
for idx, val in enumerate(row_vals):
    print(f"Col {idx+1}: {val}")
