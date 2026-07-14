import requests
import openpyxl

SUPABASE_URL = 'https://mujqmdmzloizqhglayxe.supabase.co'
SUPABASE_KEY = 'eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6Im11anFtZG16bG9penFoZ2xheXhlIiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc4MjgxNTk0OCwiZXhwIjoyMDk4MzkxOTQ4fQ.G7AmEylxPwm6TWZ3xjCcBhvhfNPYHdj0V08My0A_rEc'
headers = {
    'apikey': SUPABASE_KEY,
    'Authorization': f'Bearer {SUPABASE_KEY}'
}

# Fetch all from round_1_evaluation
url = f"{SUPABASE_URL}/rest/v1/round_1_evaluation?select=id,app_status,raw_submissions(full_name)"
r = requests.get(url, headers=headers)
db_map = {}
for row in r.json():
    raw = row.get('raw_submissions')
    if raw:
        raw_parsed = raw[0] if isinstance(raw, list) else raw
        name = raw_parsed.get('full_name')
        if name:
            db_map[name.strip().lower()] = {
                "id": row['id'],
                "status": row['app_status']
            }

python_path = r"c:\Users\Dhanush\Music\aviatorsclaude\r1_r2_side_by_side_report_2026-07-13.xlsx"
wb = openpyxl.load_workbook(python_path)
sheet = wb['Analysis']

mismatches = 0
print("Comparing DB status to Excel status:")
for r in range(7, sheet.max_row + 1):
    name = sheet.cell(row=r, column=2).value
    status_excel = sheet.cell(row=r, column=39).value # AM
    
    if name:
        name_key = str(name).strip().lower()
        if name_key in db_map:
            status_db = db_map[name_key]["status"]
            
            # Normalize status comparison
            norm_excel = str(status_excel or '').strip().lower()
            norm_db = str(status_db or '').strip().lower()
            
            if norm_excel != norm_db:
                mismatches += 1
                print(f"Mismatch: Name='{name}', DB='{status_db}', Excel='{status_excel}'")
        else:
            print(f"Name in Excel but not in DB mapping: '{name}'")

print(f"\nTotal status mismatches: {mismatches}")
