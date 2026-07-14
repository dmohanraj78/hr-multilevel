import openpyxl

file_ref = r"c:\Users\Dhanush\Music\aviatorsclaude\r1_r2_side_by_side_report_2026-07-13.xlsx"

try:
    wb = openpyxl.load_workbook(file_ref)
    sheet = wb['Analysis']
    
    # Let's inspect rows 7 to 12 for columns 39 (AM) and 46 (AT)
    print("Row inspection:")
    for r in range(7, 13):
        cell_am = sheet.cell(row=r, column=39) # Status
        cell_at = sheet.cell(row=r, column=46) # Decision
        
        print(f"Row {r}:")
        print(f"  Col 39 (Status): Value={cell_am.value}, Fill={cell_am.fill.start_color.rgb if cell_am.fill else None}, FontColor={cell_am.font.color.rgb if cell_am.font and cell_am.font.color else None}")
        print(f"  Col 46 (Decision): Value={cell_at.value}, Fill={cell_at.fill.start_color.rgb if cell_at.fill else None}, FontColor={cell_at.font.color.rgb if cell_at.font and cell_at.font.color else None}")
except Exception as e:
    print("Error:", e)
