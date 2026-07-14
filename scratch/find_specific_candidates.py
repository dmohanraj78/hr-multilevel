import openpyxl

python_path = r"c:\Users\Dhanush\Music\aviatorsclaude\r1_r2_side_by_side_report_2026-07-13.xlsx"
wb = openpyxl.load_workbook(python_path)
sheet = wb['Analysis']

# Search for candidates with tech stack containing 'n8n' or 'Qwen' or 'RAG'
print(f"Searching in {python_path}:")
for r in range(7, sheet.max_row + 1):
    name = sheet.cell(row=r, column=2).value
    tech = sheet.cell(row=r, column=43).value # AQ is column 43 (Tech stack used)
    status = sheet.cell(row=r, column=39).value # AM is R1 Status
    decision = sheet.cell(row=r, column=46).value # AT is R2 Decision
    
    if tech and any(kw in str(tech).lower() for kw in ['n8n', 'qwen', 'rag']):
        print(f"Row {r}: Name='{name}', Tech='{tech}', R1 Status='{status}', R2 Decision='{decision}'")
