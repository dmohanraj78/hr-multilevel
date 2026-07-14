import openpyxl

correct_path = r"c:\Users\Dhanush\Music\aviatorsclaude\r1_r2_side_by_side_report_2026-07-13 (4).xlsx"
python_path = r"c:\Users\Dhanush\Music\aviatorsclaude\r1_r2_side_by_side_report_2026-07-13.xlsx"

wb_correct = openpyxl.load_workbook(correct_path)
sheet_correct = wb_correct['Analysis']

wb_python = openpyxl.load_workbook(python_path)
sheet_python = wb_python['Analysis']

# Map candidate names to their row data in both sheets
def get_candidates_map(sheet):
    cand_map = {}
    for r in range(7, sheet.max_row + 1):
        name = sheet.cell(row=r, column=2).value
        if name:
            name_key = str(name).strip().lower()
            # columns AJ (36), AK (37), AL (38), AM (39)
            # wait: python openpyxl is 1-indexed, so:
            # 36: R1 Review (AJ)
            # 37: R1 Interview Priority (AK)
            # 38: To be screened by (AL)
            # 39: Status (AM)
            cand_map[name_key] = {
                "row": r,
                "review": sheet.cell(row=r, column=36).value,
                "priority": sheet.cell(row=r, column=37).value,
                "screener": sheet.cell(row=r, column=38).value,
                "status": sheet.cell(row=r, column=39).value,
                # also check column AT (46) - Decision
                "decision": sheet.cell(row=r, column=46).value,
            }
    return cand_map

map_correct = get_candidates_map(sheet_correct)
map_python = get_candidates_map(sheet_python)

print("Comparing R1 Inputs (Review, Priority, Screener, Status) and R2 Decision:")
differences = 0
for name, data_c in map_correct.items():
    if name in map_python:
        data_p = map_python[name]
        diffs = []
        if data_c["review"] != data_p["review"]:
            diffs.append(f"Review: Correct='{data_c['review']}', Python='{data_p['review']}'")
        if data_c["priority"] != data_p["priority"]:
            diffs.append(f"Priority: Correct='{data_c['priority']}', Python='{data_p['priority']}'")
        if data_c["screener"] != data_p["screener"]:
            diffs.append(f"Screener: Correct='{data_c['screener']}', Python='{data_p['screener']}'")
        if data_c["status"] != data_p["status"]:
            diffs.append(f"Status: Correct='{data_c['status']}', Python='{data_p['status']}'")
        if data_c["decision"] != data_p["decision"]:
            diffs.append(f"Decision: Correct='{data_c['decision']}', Python='{data_p['decision']}'")
            
        if diffs:
            differences += 1
            print(f"\nCandidate: {name}")
            for d in diffs:
                print(f"  {d}")
            if differences >= 20:
                print("\nToo many differences, stopping print...")
                break
                
if differences == 0:
    print("No differences found between matching candidates in R1 inputs and R2 Decision!")
else:
    print(f"\nTotal candidates with differences: {differences}")
