import openpyxl

correct_path = r"c:\Users\Dhanush\Music\aviatorsclaude\r1_r2_side_by_side_report_2026-07-13 (4).xlsx"
python_path = r"c:\Users\Dhanush\Music\aviatorsclaude\r1_r2_side_by_side_report_2026-07-13.xlsx"

try:
    wb_correct = openpyxl.load_workbook(correct_path)
    sheet_correct = wb_correct['Analysis']
    
    wb_python = openpyxl.load_workbook(python_path)
    sheet_python = wb_python['Analysis']
    
    print(f"Correct sheet rows: {sheet_correct.max_row}")
    print(f"Python sheet rows: {sheet_python.max_row}")
    
    # Let's extract names from Column 2 (B) for both sheets
    names_correct = []
    for r in range(7, sheet_correct.max_row + 1):
        val = sheet_correct.cell(row=r, column=2).value
        if val:
            names_correct.append(str(val).strip().lower())
            
    names_python = []
    for r in range(7, sheet_python.max_row + 1):
        val = sheet_python.cell(row=r, column=2).value
        if val:
            names_python.append(str(val).strip().lower())
            
    print(f"Correct sheet unique names count: {len(set(names_correct))}")
    print(f"Python sheet unique names count: {len(set(names_python))}")
    
    missing_in_python = set(names_correct) - set(names_python)
    print(f"\nNames in correct report but missing in Python report ({len(missing_in_python)}):")
    for name in sorted(list(missing_in_python))[:15]:
        print(f"  {name}")
        
    missing_in_correct = set(names_python) - set(names_correct)
    print(f"\nNames in Python report but missing in correct report ({len(missing_in_correct)}):")
    for name in sorted(list(missing_in_correct))[:15]:
        print(f"  {name}")
        
except Exception as e:
    print("Error:", e)
