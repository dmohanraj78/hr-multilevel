import openpyxl

python_path = r"c:\Users\Dhanush\Music\aviatorsclaude\r1_r2_side_by_side_report_2026-07-13.xlsx"
wb = openpyxl.load_workbook(python_path)
sheet = wb['Analysis']

print(f"Total rows in generated report: {sheet.max_row}")
print("\nFirst 20 candidate rows details (Columns B, AJ, AK, AL, AM, AT):")
print(f"{'Name':<35} | {'Review (AJ)':<25} | {'Priority (AK)':<13} | {'Screener (AL)':<13} | {'Status (AM)':<10} | {'Decision (AT)':<10}")
print("-" * 120)
for r in range(7, 27):
    name = sheet.cell(row=r, column=2).value
    review = sheet.cell(row=r, column=36).value
    priority = sheet.cell(row=r, column=37).value
    screener = sheet.cell(row=r, column=38).value
    status = sheet.cell(row=r, column=39).value
    decision = sheet.cell(row=r, column=46).value
    print(f"{str(name or ''):<35} | {str(review or ''):<25} | {str(priority or ''):<13} | {str(screener or ''):<13} | {str(status or ''):<10} | {str(decision or ''):<10}")
