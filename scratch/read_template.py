import openpyxl

wb = openpyxl.load_workbook('master-portal/public/template.xlsx')
sheet = wb['Analysis']

print("Row 3 values:")
for col in range(1, 16):
    print(sheet.cell(row=3, column=col).value, end=" | ")
print("\nRow 4 values:")
for col in range(1, 16):
    print(sheet.cell(row=4, column=col).value, end=" | ")
print()
