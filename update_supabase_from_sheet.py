import requests
import openpyxl

SUPABASE_URL = "https://mujqmdmzloizqhglayxe.supabase.co"
SUPABASE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6Im11anFtZG16bG9penFoZ2xheXhlIiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc4MjgxNTk0OCwiZXhwIjoyMDk4MzkxOTQ4fQ.G7AmEylxPwm6TWZ3xjCcBhvhfNPYHdj0V08My0A_rEc"

headers = {
    "apikey": SUPABASE_KEY,
    "Authorization": f"Bearer {SUPABASE_KEY}",
    "Content-Type": "application/json"
}

# Fetch R1 reviews + raw submissions
print("Fetching candidate data from Supabase...")
r1_url = f"{SUPABASE_URL}/rest/v1/round_1_evaluation?select=*,raw_submissions(*)"
resp_r1 = requests.get(r1_url, headers=headers)
r1_list = resp_r1.json()

unprocessed = []
for r1 in r1_list:
    if r1.get("total") is None and r1.get("tier") is None:
        raw = r1.get("raw_submissions")
        if raw:
            raw_parsed = raw[0] if isinstance(raw, list) else raw
            unprocessed.append({
                "id": r1["id"],
                "full_name": raw_parsed.get("full_name"),
                "email": raw_parsed.get("email")
            })

print(f"Total unprocessed in Supabase: {len(unprocessed)}")

wb = openpyxl.load_workbook(r"c:\Users\Dhanush\Music\aviatorsclaude\Untitled spreadsheet (1).xlsx")
ws = wb.active

# Load spreadsheet data
sheet_data = []
for r in range(2, ws.max_row + 1):
    row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
    sheet_data.append({
        "row": r,
        "name": str(row_vals[1] or "").strip(),
        "email": str(row_vals[27] or "").strip(),
        "vals": row_vals
    })

print("\n--- Updating matched unprocessed candidates in Supabase ---")
updated_count = 0
for u in unprocessed:
    matched_row = None
    for sd in sheet_data:
        # Match by email first
        if u["email"] and sd["email"] and u["email"].lower() == sd["email"].lower():
            matched_row = sd
            break
        # Match by name
        if u["full_name"] and sd["name"] and u["full_name"].lower() == sd["name"].lower():
            matched_row = sd
            break

    if matched_row:
        vals = matched_row["vals"]
        total = float(vals[6]) if vals[6] is not None else None
        
        # Calculate review_cat
        review_cat = None
        if total is not None:
            if total >= 16:
                review_cat = "Strong"
            elif total >= 12:
                review_cat = "Good"
            elif total >= 8:
                review_cat = "Need Clarity"
            else:
                review_cat = "Invalid"

        # Determine app_status
        col35 = str(vals[34] or "").strip()
        app_status = "Pending"
        if col35.lower() == "yes":
            app_status = "Yes"
        elif col35.lower() == "reject":
            app_status = "Reject"
        elif col35.lower() == "maybe":
            app_status = "Maybe"
        elif col35.lower() == "duplicate":
            app_status = "Duplicate"

        patch_data = {
            "gender": vals[2],
            "cat": vals[3],
            "graduation": str(vals[4]) if vals[4] is not None else None,
            "tier": vals[5],
            "total": total,
            "edu": float(vals[7]) if vals[7] is not None else None,
            "exp": float(vals[8]) if vals[8] is not None else None,
            "proj": float(vals[9]) if vals[9] is not None else None,
            "substance": float(vals[10]) if vals[10] is not None else None,
            "deploy": float(vals[11]) if vals[11] is not None else None,
            "artifact": float(vals[12]) if vals[12] is not None else None,
            "skills": float(vals[13]) if vals[13] is not None else None,
            "domain": vals[14],
            "degree": vals[15],
            "stream": vals[16],
            "college": vals[17],
            "location": vals[18],
            "ai_proj": float(vals[19]) if vals[19] is not None else None,
            "fs_proj": float(vals[20]) if vals[20] is not None else None,
            "intern_mo": float(vals[21]) if vals[21] is not None else None,
            "co_tier": float(vals[22]) if vals[22] is not None else None,
            "deploy_stage": vals[23],
            "num_skills": int(vals[24]) if vals[24] is not None else None,
            "claude_lvl": vals[25],
            "aiml_exp": vals[26],
            "demo_explanation": vals[31],
            "demo_review_notes_ai": vals[32],
            "review_comments": vals[33],
            "review_cat": review_cat,
            "app_status": app_status
        }

        # Send PATCH request to Supabase
        upd_url = f"{SUPABASE_URL}/rest/v1/round_1_evaluation?id=eq.{u['id']}"
        res = requests.patch(upd_url, headers=headers, json=patch_data)
        if res.status_code in [200, 204]:
            print(f"Successfully updated '{u['full_name']}' (ID {u['id']}) from Row {matched_row['row']}! (Total: {total}, Tier: {vals[5]}, Status: {app_status})")
            updated_count += 1
        else:
            print(f"Failed to update ID {u['id']}: {res.status_code} - {res.text}")
    else:
        print(f"Skipping '{u['full_name']}' (ID {u['id']}) — no matching evaluation data found in sheet.")

print(f"\nCompleted! Updated {updated_count} candidates in Supabase.")
