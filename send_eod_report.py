import os
import sys
import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from datetime import datetime

# Load .env file manually to avoid dependency issues if python-dotenv is not installed
def load_dotenv():
    env_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env")
    if os.path.exists(env_path):
        with open(env_path, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith("#") and "=" in line:
                    key, val = line.split("=", 1)
                    os.environ[key.strip()] = val.strip().strip('"').strip("'")

load_dotenv()

# Enforce exactly 9 AM local Chicago time (CDT/CST) in GitHub Actions runs
def check_chicago_hour():
    if not os.getenv("GITHUB_ACTIONS"):
        return
    try:
        from zoneinfo import ZoneInfo
        chicago_tz = ZoneInfo("America/Chicago")
        now_chicago = datetime.now(chicago_tz)
        hour = now_chicago.hour
        print(f"Checking Chicago hour: {now_chicago.strftime('%Y-%m-%d %I:%M %p')}")
        if hour != 9:
            print(f"Chicago hour is {hour}, not 9 AM. Skipping daily run.")
            sys.exit(0)
    except Exception as e:
        print(f"Error checking Chicago timezone hour: {e}. Proceeding anyway.")

check_chicago_hour()

# Configuration
SUPABASE_URL = os.getenv("SUPABASE_URL", "https://mujqmdmzloizqhglayxe.supabase.co")
SUPABASE_KEY = os.getenv("SUPABASE_KEY", "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6Im11anFtZG16bG9penFoZ2xheXhlIiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc4MjgxNTk0OCwiZXhwIjoyMDk4MzkxOTQ4fQ.G7AmEylxPwm6TWZ3xjCcBhvhfNPYHdj0V08My0A_rEc")
SMTP_HOST = os.getenv("SMTP_HOST", "smtp.gmail.com")
SMTP_PORT = int(os.getenv("SMTP_PORT", "587"))
SMTP_USER = os.getenv("SMTP_USER", "")
SMTP_PASSWORD = os.getenv("SMTP_PASSWORD", "")
EMAIL_FROM = os.getenv("EMAIL_FROM", "")
EMAIL_TO = os.getenv("EMAIL_TO", "mgoel@mondee.com")  # Default recipient

def get_median(lst):
    n = len(lst)
    if n == 0:
        return 0
    s_lst = sorted(lst)
    mid = n // 2
    if n % 2 == 1:
        return s_lst[mid]
    else:
        return (s_lst[mid - 1] + s_lst[mid]) / 2.0

def build_excel_report():
    print("Fetching candidate data from Supabase...")
    headers = {
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}"
    }

    # Fetch R1 reviews + raw submissions
    r1_url = f"{SUPABASE_URL}/rest/v1/round_1_evaluation?select=*,raw_submissions(*)"
    resp_r1 = requests.get(r1_url, headers=headers)
    if resp_r1.status_code != 200:
        print(f"Error fetching R1: {resp_r1.text}")
        sys.exit(1)
    r1_list = resp_r1.json()

    # Fetch R2 reviews
    r2_url = f"{SUPABASE_URL}/rest/v1/round_2_evaluation?select=*"
    resp_r2 = requests.get(r2_url, headers=headers)
    if resp_r2.status_code != 200:
        print(f"Error fetching R2: {resp_r2.text}")
        sys.exit(1)
    r2_list = resp_r2.json()

    # Map R2 by ID
    r2_map = {item["id"]: item for item in r2_list}

    # Filter evaluated candidates (having R1 record)
    candidates = []
    for r1 in r1_list:
        raw = r1.get("raw_submissions")
        if not raw:
            continue
        # Every round_1_evaluation row counts as an applicant — rows the import
        # pipeline created that are not yet AI-scored surface as Pending, so the
        # report total always equals count(round_1_evaluation).
        raw_parsed = raw[0] if isinstance(raw, list) else raw
        candidates.append({
            "c": raw_parsed,
            "r1": r1,
            "r2": r2_map.get(r1["id"], {})
        })

    # Sort order (primary = R1 Status so reviewed people stay together):
    #   1. R1 Yes (reviewed & moved to Round 2) — within this block the
    #      R2-vetted candidates come first (R2 Yes -> Maybe -> No -> vetting
    #      started without decision), untouched/null R2 rows sink to its end
    #   2. R1 Maybe
    #   3. R1 Rejected
    #   4. Pending / everything else last (the mostly-empty rows)
    def sort_key(cand):
        r1_status = str(cand["r1"].get("app_status") or "").strip().lower()
        r2 = cand["r2"]
        # _draft decisions are unfinished reviews — treat as no decision yet
        # (matches the dashboard, which only counts finalized decisions)
        r2_decision_raw = str(r2.get("moved_to_round_3") or "").strip()
        r2_decision = "" if r2_decision_raw.endswith("_draft") else r2_decision_raw.lower()
        score = float(cand["r1"].get("total") or 0)

        if r1_status == "yes":
            group = 0
        elif r1_status == "maybe":
            group = 1
        elif r1_status in ["no", "rejected", "invalid", "reject"]:
            group = 2
        else:
            group = 3  # Pending & rest

        if r2_decision in ["yes", "promoted"]:
            sub = 0
        elif r2_decision == "maybe":
            sub = 1
        elif r2_decision in ["no", "declined"]:
            sub = 2
        elif r2.get("id") is not None:
            sub = 3  # vetting started, no decision yet
        else:
            sub = 4  # no R2 record at all

        return (group, sub, -score)

    candidates.sort(key=sort_key)

    # Initialize Workbook from template if exists
    template_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Copy of AI_Builder_Intern_v9.xlsx")
    template_loaded = False
    
    if os.path.exists(template_path):
        print(f"Loading template from {template_path}...")
        wb = openpyxl.load_workbook(template_path)
        ws = wb["Analysis"]
        template_loaded = True
        # Clear existing candidate rows in Analysis sheet (starting from Row 7 onwards)
        if ws.max_row >= 7:
            ws.delete_rows(7, ws.max_row - 6)
    else:
        print("Template file not found. Creating blank workbook.")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Analysis"
        ws.views.sheetView[0].showGridLines = True

        # Row 1: Title
        ws.row_dimensions[1].height = 28
        ws["A1"] = "Scored Candidates — Analysis"
        ws["A1"].font = Font(name="Segoe UI", size=16, bold=True, color="1F3864")

        # Row 2: Subtitle
        ws.row_dimensions[2].height = 42
        ws["A2"] = "Demo Review Notes (AG): DARK GREEN = Industry problem + advanced AI + complex + not a student project · LIGHT GREEN = Function problem + advanced AI + complex + not common student project · PINK = unclear/unassessable · RED = everything else.   Demo cell (AE) RED = invalid/missing demo link."
        ws["A2"].font = Font(name="Segoe UI", size=9.5, italic=True, color="595959")
        ws["A2"].alignment = Alignment(wrap_text=True, vertical="center")

    # Statistics Calculation
    total_applicants = len(candidates)
    t1 = 0
    t1_plus = 0
    t2 = 0
    t2_plus = 0
    t3 = 0
    t4 = 0
    scores = []

    for cand in candidates:
        r1 = cand["r1"]
        tier = str(r1.get("tier") or "").strip().upper()
        if tier in ["T1", "TIER 1"]:
            t1 += 1
        elif tier in ["T1+", "TIER 1+"]:
            t1_plus += 1
        elif tier in ["T2", "TIER 2"]:
            t2 += 1
        elif tier in ["T2+", "TIER 2+"]:
            t2_plus += 1
        elif tier in ["T3", "TIER 3"]:
            t3 += 1
        elif tier in ["T4", "TIER 4"]:
            t4 += 1

        # Only scored rows feed the average/median — unprocessed Pending rows
        # have no total yet and would drag the stats down as zeros
        if r1.get("total") is not None:
            scores.append(float(r1.get("total") or 0))

    avg_score = sum(scores) / len(scores) if scores else 0
    top_score = max(scores) if scores else 0
    median_score = get_median(scores)

    # Write Row 3 & 4 (Stats Table)
    ws.row_dimensions[3].height = 20
    ws.row_dimensions[4].height = 18

    # Funnel reconciliation counts — mirror the Supabase tables directly so the
    # sheet's numbers can be sanity-checked against the dashboard
    def _r1_status(cand):
        status_val = str(cand["r1"].get("app_status") or "").strip().lower()
        if not status_val or status_val in ["", "-", "none"]:
            return "pending"
        return status_val
    f_moved_to_r2 = sum(1 for cand in candidates if _r1_status(cand) == "yes")
    f_r2_evaluated = sum(1 for cand in candidates if cand["r2"].get("id") is not None)
    f_rejected = sum(1 for cand in candidates if _r1_status(cand) in ["no", "rejected", "invalid", "reject"])
    f_maybe = sum(1 for cand in candidates if _r1_status(cand) == "maybe")
    f_pending = total_applicants - f_moved_to_r2 - f_rejected - f_maybe

    # Generation stamp so a stale report is immediately recognisable
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M")
    stats_vals = [total_applicants, t1, t1_plus, t2, t2_plus, t3, t4, round(avg_score, 1), top_score, median_score, f_moved_to_r2, f_r2_evaluated, f_rejected, f_pending, generated_at]
    stats_labels = ["Applicants", "Tier 1", "Tier 1+", "Tier 2", "Tier 2+", "Tier 3", "Tier 4", "Avg Total", "Top Score", "Median", "Moved to R2", "R2 Evaluated", "Rejected", "Pending", "Generated"]

    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    light_blue_fill = PatternFill(start_color="F2F5FB", end_color="F2F5FB", fill_type="solid")

    for i in range(len(stats_vals)):
        col_letter = get_column_letter(i + 1)
        
        # Row 3 Value
        cell_val = ws[f"{col_letter}3"]
        cell_val.value = stats_vals[i]
        cell_val.font = Font(name="Segoe UI", size=10, bold=True, color="1F3864")
        cell_val.fill = light_blue_fill
        cell_val.alignment = Alignment(horizontal="center", vertical="center")
        cell_val.border = thin_border

        # Row 4 Label
        cell_lbl = ws[f"{col_letter}4"]
        cell_lbl.value = stats_labels[i]
        cell_lbl.font = Font(name="Segoe UI", size=9, color="595959")
        cell_lbl.fill = light_blue_fill
        cell_lbl.alignment = Alignment(horizontal="center", vertical="center")
        cell_lbl.border = thin_border

    # Overwrite Raw Data sheet if it exists
    if "Raw Data" in wb.sheetnames:
        ws_raw = wb["Raw Data"]
        raw_row_idx = 2
        for cand in candidates:
            c = cand["c"]
            r1 = cand["r1"]
            row_data = [
                c.get("id"),
                c.get("submission_date"),
                c.get("full_name"),
                c.get("email"),
                c.get("phone"),
                c.get("linkedin"),
                c.get("location"),
                c.get("preferred_start_date"),
                c.get("education_level"),
                c.get("ug_university"),
                c.get("masters_university"),
                c.get("course_major"),
                c.get("degree_status"),
                c.get("completion_year"),
                c.get("programming_languages"),
                c.get("aiml_experience") or r1.get("aiml_exp") or '',
                c.get("claude_ecosystem") or r1.get("claude_lvl") or '',
                c.get("skill_python"),
                c.get("skill_deep_learning"),
                c.get("skill_nlp"),
                c.get("skill_computer_vision"),
                c.get("skill_reinforcement_learning"),
                c.get("skill_multimodal_ai"),
                c.get("skill_finetuning_lora_peft"),
                c.get("skill_llm_orchestration"),
                c.get("skill_agent_fundamentals"),
                c.get("skill_mcp"),
                c.get("skill_embeddings_vector_rag"),
                c.get("skill_reasoning_models"),
                c.get("skill_evals"),
                c.get("skill_ai_coding_tools"),
                c.get("skill_rag"),
                c.get("demo_explanation") or c.get("current_project") or '',
                c.get("project_state") or r1.get("deploy_stage") or '',
                c.get("project_category") or r1.get("domain") or '',
                c.get("demo_link"),
                c.get("github_url"),
                c.get("preferred_industry") or '',
                c.get("open_to_onsite") or '',
                c.get("open_to_travel") or '',
                c.get("anything_else") or '',
                c.get("applied_role") or '',
                c.get("job_id") or '',
                c.get("rubric_id") or '',
                c.get("screening_type") or '',
                c.get("resume_drive_url"),
                c.get("resume_filename") or '',
                c.get("Analysis_status") or ''
            ]
            for col_idx, val in enumerate(row_data):
                ws_raw.cell(row=raw_row_idx, column=col_idx + 1).value = val
            raw_row_idx += 1
            
        # Delete any trailing/leftover rows from template
        if ws_raw.max_row >= raw_row_idx:
            ws_raw.delete_rows(raw_row_idx, ws_raw.max_row - raw_row_idx + 1)

    grid_border = Border(
        left=Side(style='thin', color='E0E0E0'),
        right=Side(style='thin', color='E0E0E0'),
        top=Side(style='thin', color='E0E0E0'),
        bottom=Side(style='thin', color='E0E0E0')
    )

    # Row 5: Merged Headers for Round 1 & Round 2 Inputs
    ws.row_dimensions[5].height = 18
    # Unmerge ANY existing merged range that touches row 5 so we can re-merge cleanly
    for r in list(ws.merged_cells.ranges):
        if r.min_row <= 5 <= r.max_row:
            try:
                ws.unmerge_cells(r.coord)
            except Exception:
                pass

    # Three labelled sections: candidate profile (A-AI), the columns the R1
    # reviewer fills (AJ-AM, ending at Status), and the R2 inputs (AN-AU) —
    # so each label sits directly above its own columns.
    row5_sections = [
        ('A5:AI5', 'A5', 'Candidate Analysed Details', '1F3864'),
        ('AJ5:AM5', 'AJ5', 'Round 1 Inputs', '2F5597'),
        ('AN5:AU5', 'AN5', 'Round 2 Inputs', '0070C0'),
    ]
    for rng, anchor, label, color in row5_sections:
        ws.merge_cells(rng)
        cell = ws[anchor]
        cell.value = label
        cell.font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    if not template_loaded:
        # Row 6: Column Headers
        headers = [
            "Rank", "Name", "Gender", "Cat", "Graduation", "Tier", "Total", "Edu", "Exp", "Proj", 
            "Substance", "Deploy", "Artifact", "Skills", "Domain", "Degree", "Stream", "College", "F_college", "F_University", 
            "Location", "AI Proj", "FS Proj", "Intern Mo", "Co.Tier", "Deploy Stage", "#Skills", "Claude Lvl", "AI/ML Exp", "Email", 
            "Résumé", "GitHub", "Demo", "Demo Explanation (their project)", "Demo Review Notes (AI)", "R1 Review", "R1 Interview Priority", "To be screened by", 
            "Status", "Earliest date they can start the internship", "Any concerns / restrictions (with college commitment, personal, others)", "Technical depth of demo / product", "Tech stack used", "Problem-solution fit", "Areas like latency, cost, security, etc been considered", "Decision", "Reason for decision (detailed notes)"
        ]

        ws.row_dimensions[6].height = 24
        dark_blue_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
        mid_blue_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
        blue_fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")

        for idx, h in enumerate(headers):
            cell = ws.cell(row=6, column=idx + 1)
            cell.value = h
            cell.font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
            # Match row-5 sections: profile (0-34), R1 inputs (35-38), R2 inputs (39+)
            cell.fill = blue_fill if idx >= 39 else (mid_blue_fill if idx >= 35 else dark_blue_fill)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = grid_border

    # Fills for row data highlighting
    green_fill = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")
    green_font = Font(name="Segoe UI", size=9.5, color="385723", bold=True)
    
    amber_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    amber_font = Font(name="Segoe UI", size=9.5, color="856404", bold=True)
    
    red_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
    red_font = Font(name="Segoe UI", size=9.5, color="721C24", bold=True)

    # Write Candidates starting at Row 7
    for index, cand in enumerate(candidates):
        row_number = index + 7
        ws.row_dimensions[row_number].height = 20
        c = cand["c"]
        r1 = cand["r1"]
        r2 = cand["r2"]

        raw_solves = r2.get("solves_business_problem") or ""
        
        # Safe string checks
        contact_status = r2.get("contact_status") or ""
        if not contact_status:
            if "Contact: " in raw_solves:
                contact_status = raw_solves.split("Contact: ")[1].split(" | ")[0]
            elif raw_solves in ["Yet to Speak", "Spoke", "Scheduled", "No response"]:
                contact_status = raw_solves

        problem_fit = r2.get("problem_fit") or ""
        if not problem_fit:
            if "Fit: " in raw_solves:
                problem_fit = raw_solves.split("Fit: ")[1]
            elif raw_solves in ["Yes", "Maybe", "No"]:
                problem_fit = raw_solves

        raw_depth = r2.get("product_depth") or ""
        tech_depth = r2.get("tech_depth") or ""
        if not tech_depth:
            if "Depth: " in raw_depth:
                tech_depth = raw_depth.split("Depth: ")[1].split(" | ")[0]
            elif raw_depth in ["High", "Medium", "Low", "None"]:
                tech_depth = raw_depth

        latency = r2.get("latency_considerations") or ""
        if not latency:
            if "Latency: " in raw_depth:
                latency = raw_depth.split("Latency: ")[1]
            elif raw_depth not in ["High", "Medium", "Low", "None"]:
                latency = raw_depth

        # R1 Status
        r1_status_val = r1.get("app_status")
        if not r1_status_val or str(r1_status_val).strip() in ["", "-", "None"]:
            r1_status_val = "Pending"

        # R2 Inputs formatting
        r2_has_id = r2.get("id") is not None
        
        def r2_clean(val):
            if not r2_has_id:
                return "-"
            if val is None or str(val).strip() in ["", "-", "None"]:
                return "NA"
            return val

        r2_start = r2_clean(r2.get("when_can_they_start"))
        r2_complexity = r2_clean(r2.get("complexity"))
        r2_tech_depth = r2_clean(tech_depth)
        r2_stack = r2_clean(r2.get("tech_stack"))
        r2_fit = r2_clean(problem_fit)
        r2_latency = r2_clean(latency)
        r2_comment = r2_clean(r2.get("demo_review_comment"))

        # R2 Decision logic
        if r2_has_id:
            r2_decision_val = r2.get("moved_to_round_3") or "Pending"
            # _draft = reviewer saved but did not finish — not a decision yet
            if str(r2_decision_val).strip().endswith("_draft"):
                r2_decision_val = "Pending"
            if not r2_decision_val or str(r2_decision_val).strip() in ["", "-", "None"]:
                r2_decision_val = "Pending"
        elif str(r1.get("app_status") or "").strip().lower() in ["no", "rejected", "invalid", "reject"]:
            r2_decision_val = "Rejected (R1)"
        else:
            r2_decision_val = "Pending"

        row_data = [
            index + 1,  # Rank
            c.get("full_name"),
            r1.get("gender") or "-",
            r1.get("cat") or "-",
            r1.get("graduation") or "-",
            r1.get("tier") or "-",
            float(r1.get("total") or 0),
            float(r1.get("edu") or 0),
            float(r1.get("exp") or 0),
            float(r1.get("proj") or 0),
            float(r1.get("substance") or 0),
            float(r1.get("deploy") or 0),
            float(r1.get("artifact") or 0),
            float(r1.get("skills") or 0),
            r1.get("domain") or "-",
            r1.get("degree") or "-",
            r1.get("stream") or "-",
            c.get("ug_university") or "-",
            r1.get("college") or "-",
            c.get("ug_university") or "-",
            r1.get("location") or "-",
            float(r1.get("ai_proj") or 0),
            float(r1.get("fs_proj") or 0),
            float(r1.get("intern_mo") or 0),
            float(r1.get("co_tier") or 0),
            r1.get("deploy_stage") or "-",
            float(r1.get("num_skills") or 0),
            r1.get("claude_lvl") or "-",
            r1.get("aiml_exp") or "-",
            c.get("email") or "-",
            c.get("resume_drive_url") or "-",
            c.get("github_url") or "-",
            c.get("demo_link") or "-",
            c.get("demo_explanation") or c.get("current_project") or "-",
            r1.get("demo_review_notes_ai") or "-",
            r1.get("review_comments") or "-",
            r1.get("r1_interview_priority") or "-",
            r1.get("eval_group") or "-",
            r1_status_val,
            r2_start,
            r2_complexity,
            r2_tech_depth,
            r2_stack,
            r2_fit,
            r2_latency,
            r2_decision_val,
            # Tier column is removed!
            r2_comment
        ]

        # Column indices (0-based) for decision cells that get color-coded:
        # idx 38 = R1 "Status" (col 39), idx 45 = R2 "Decision" (col 46)
        DECISION_COLS = {38, 45}

        for idx, val in enumerate(row_data):
            cell = ws.cell(row=row_number, column=idx + 1)
            cell.value = val
            cell.font = Font(name="Segoe UI", size=9.5)
            cell.alignment = Alignment(vertical="center", horizontal="center" if isinstance(val, (int, float)) else "left")
            cell.border = grid_border

            # Color-code ONLY the R1 Status and R2 Decision columns
            if idx in DECISION_COLS:
                val_str = str(val or "").strip().lower()
                if val_str in ["yes", "promoted", "approved"]:
                    cell.fill = green_fill
                    cell.font = green_font
                elif val_str in ["maybe"]:
                    cell.fill = amber_fill
                    cell.font = amber_font
                elif val_str in ["no", "invalid", "declined"] or val_str.startswith("reject"):
                    cell.fill = red_fill
                    cell.font = red_font
                else:
                    cell.fill = PatternFill(fill_type=None)
                    cell.font = Font(name="Segoe UI", size=9.5, color="000000")

    # Set Column Widths only if template not loaded
    if not template_loaded:
        widths = [
            6, 25, 10, 8, 12, 10, 8, 8, 8, 8, 8, 8, 8, 8, 20, 15, 20, 30, 30, 25,
            18, 10, 10, 10, 10, 18, 10, 15, 15, 30, 25, 25, 25, 35, 35, 30, 18, 18,
            12, 18, 25, 18, 25, 15, 20, 12, 35
        ]
        for idx, w in enumerate(widths):
            ws.column_dimensions[get_column_letter(idx + 1)].width = w

    # Unhide all rows in the sheet to prevent template from hiding data rows
    for r in range(1, ws.max_row + 1):
        ws.row_dimensions[r].hidden = False

    date_str = datetime.now().strftime("%Y-%m-%d")
    output_filename = f"R1_R2_EOD_REPORT_{date_str}.xlsx"
    wb.save(output_filename)
    print(f"Report saved successfully as {output_filename}")
    return output_filename

def _fetch_all_rows(table, select):
    req_headers = {
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
    }
    rows, offset = [], 0
    while True:
        r = requests.get(
            f"{SUPABASE_URL}/rest/v1/{table}?select={select}&limit=1000&offset={offset}",
            headers=req_headers,
        )
        r.raise_for_status()
        page = r.json()
        rows += page
        if len(page) < 1000:
            return rows
        offset += 1000


def build_funnel_summary():
    """Compute the funnel narrative from live Supabase data at send time."""
    raw = _fetch_all_rows("raw_submissions", "id,Analysis_status")
    r1 = _fetch_all_rows("round_1_evaluation", "id,tier,app_status,review_comments,eval_group")
    r2 = _fetch_all_rows("round_2_evaluation", "id,moved_to_round_3")
    r3 = _fetch_all_rows("round_3_evaluation", "id,verdict")

    r1_ids = {x["id"] for x in r1}
    no_r1 = [x for x in raw if x["id"] not in r1_ids]
    duplicates = sum(1 for x in no_r1 if (x.get("Analysis_status") or "") == "Completed")
    pending_analysis = len(no_r1) - duplicates
    received = len(raw)
    unique = received - duplicates
    reviewed = len(r1)

    def status(x):
        return (x.get("app_status") or "Pending").strip()

    def touched(x):
        return status(x) != "Pending" or bool((x.get("review_comments") or "").strip())

    def tier_of(x):
        return (x.get("tier") or "").strip()

    TOP = {"Tier 1", "Tier 1+", "Tier 2", "Tier 2+"}
    LOW = {"Tier 3", "Tier 4"}
    top = [x for x in r1 if tier_of(x) in TOP]
    low = [x for x in r1 if tier_of(x) in LOW]

    top_touched = [x for x in top if touched(x)]
    top_yes = sum(1 for x in top_touched if status(x) == "Yes")
    top_no = sum(1 for x in top_touched if status(x) in ("Reject", "No", "Rejected", "Invalid"))
    top_pend = len(top_touched) - top_yes - top_no

    low_touched = [x for x in low if touched(x)]
    low_yes = sum(1 for x in low_touched if status(x) == "Yes")
    low_pend = len(low) - len(low_touched)

    r1_yes = sum(1 for x in r1 if status(x) == "Yes")
    yes_rows = [x for x in r1 if status(x) == "Yes"]
    assigned = sum(1 for x in yes_rows if (x.get("eval_group") or "").strip() not in ("", "None", "Unassigned"))
    unassigned = r1_yes - assigned

    def decision(x):
        d = (x.get("moved_to_round_3") or "").strip()
        return "" if d.endswith("_draft") else d  # drafts are not decisions

    r2_yes = sum(1 for x in r2 if decision(x) == "Yes")
    r2_maybe = sum(1 for x in r2 if decision(x) == "Maybe")
    r2_no = sum(1 for x in r2 if decision(x) == "No")
    r2_declined = sum(1 for x in r2 if decision(x) == "Declined")
    finalized = r2_yes + r2_maybe + r2_no + r2_declined
    in_progress = len(r2) - finalized
    promoted = r2_yes + r2_maybe

    hired = sum(1 for x in r3 if (x.get("verdict") or "") == "Yes")
    rejected_r3 = sum(1 for x in r3 if (x.get("verdict") or "") == "No")
    r3_pending = promoted - hired - rejected_r3

    # reviews still open = assigned candidates without a finalized decision
    # (covers both saved drafts and reviews not yet started)
    open_reviews = max(assigned - finalized, 0)
    progress_note = (
        f"{finalized} reviews are finalized and {open_reviews} are pending (in progress or not started)"
        if open_reviews
        else f"all {finalized} reviews are finalized"
    )

    return f"""We received {received} applications. {duplicates} were duplicates.
From a total of {unique} unique applications, {reviewed} were reviewed and {pending_analysis} are pending review.

Round 1:

Out of the {reviewed} reviewed, {len(top)} applicants qualified to Tier 1, Tier 1+, Tier 2 and Tier 2+. {len(top_touched)} applications have been manually reviewed and comments have been marked in Round 1.
Out of that {len(top_touched)} applications, {top_yes} got yes, {top_no} got no and {top_pend} are pending review.
Out of the {reviewed} reviewed, {len(low)} applicants qualified to Tier 3 and Tier 4. {len(low_touched)} were reviewed - out of that {low_yes} are shortlisted and {low_pend} are pending manual review.

Total of {r1_yes} out of {reviewed} applicants moved from Round 1 to Round 2.

Round 2:

Out of {r1_yes} candidates, {assigned} candidates were assigned to technical reviewers and {unassigned} applicants are yet to be assigned.
Out of the {assigned} assigned candidates, {progress_note}.
Out of the {finalized} finalized reviews, technical reviewers said yes for {r2_yes} candidates, maybe for {r2_maybe}, no for {r2_no}{f" and declined for {r2_declined}" if r2_declined else ""}.

Total of {promoted} candidates moved from Round 2 to Round 3.

Round 3:

{r3_pending} candidates pending for review. {hired} candidates hired.
{rejected_r3} candidates rejected."""


def send_email(filename, summary=""):
    if not SMTP_USER or not SMTP_PASSWORD or not EMAIL_TO:
        print("Error: SMTP details or recipient address not configured in .env. Skipping email.")
        return False

    print(f"Sending email to {EMAIL_TO}...")
    msg = MIMEMultipart()
    msg['From'] = EMAIL_FROM or SMTP_USER
    msg['To'] = EMAIL_TO
    date_str = datetime.now().strftime("%Y-%m-%d")
    msg['Subject'] = f"Aviators R1 & R2 EOD Report - {date_str}"

    body = f"""Hello Mayank,

Aviators R1 and R2 EOD report of {date_str} is attached.

{summary}

Best Regards,
Aviators Recruitment Bot"""

    msg.attach(MIMEText(body, 'plain'))

    try:
        with open(filename, "rb") as attachment:
            part = MIMEBase('application', 'octet-stream')
            part.set_payload(attachment.read())
            encoders.encode_base64(part)
            part.add_header(
                'Content-Disposition',
                f'attachment; filename= {filename}',
            )
            msg.attach(part)
            
            server = smtplib.SMTP(SMTP_HOST, SMTP_PORT)
            server.starttls()
            server.login(SMTP_USER, SMTP_PASSWORD)
            text = msg.as_string()
            server.sendmail(msg['From'], msg['To'], text)
            server.quit()
            print("Email sent successfully!")
            return True
    except Exception as e:
        print(f"Failed to send email: {e}")
        return False

if __name__ == "__main__":
    report_file = build_excel_report()
    try:
        funnel_summary = build_funnel_summary()
        print("--- Funnel summary ---")
        print(funnel_summary)
    except Exception as e:
        print(f"Failed to build funnel summary (sending without it): {e}")
        funnel_summary = ""
    sent = send_email(report_file, funnel_summary)
    # Fail the scheduled run visibly (e.g. GitHub Actions red X) if the email
    # could not be delivered, instead of silently succeeding.
    if not sent and os.getenv("GITHUB_ACTIONS"):
        sys.exit(1)
