import openpyxl

file_ref = r"c:\Users\Dhanush\Music\aviatorsclaude\r1_r2_side_by_side_report_2026-07-13.xlsx"

try:
    wb = openpyxl.load_workbook(file_ref, data_only=True)
    sheet = wb['Analysis']
    headers = [cell.value for cell in list(sheet.iter_rows(min_row=6, max_row=6))[0]]
    print("Columns:")
    for idx, h in enumerate(headers):
        print(f"  {idx}: {h}")
        
    print("\nRow 7 values:")
    vals = [cell.value for cell in list(sheet.iter_rows(min_row=7, max_row=7))[0]]
    for idx, h in enumerate(headers):
        print(f"  {h} ({idx}): {vals[idx] if idx < len(vals) else 'OUT OF RANGE'}")
except Exception as e:
    print("Error:", e)
